"""
Export Final DKP Template

This script exports player data to an Excel file where you can:
1. See all players with their fight period kills (T4/T5)
2. Enter T4 deaths and T5 deaths manually
3. See calculated DKP scores automatically

DKP Formula:
- T4 Kills × 1
- T5 Kills × 2
- T4 Deaths × 4
- T5 Deaths × 8

Usage:
    python export_dkp_template.py

The script will connect to your MongoDB and export the active season's data.
"""

import asyncio
import os
import sys
from datetime import datetime

# Add the parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# DKP Weights
T4_KILL_WEIGHT = 1
T5_KILL_WEIGHT = 2
T4_DEATH_WEIGHT = 4
T5_DEATH_WEIGHT = 8


async def get_player_data():
    """Fetch player data from MongoDB."""
    mongodb_url = os.getenv("MONGODB_URL")
    database_name = os.getenv("DATABASE_NAME", "kvk_tracker")

    if not mongodb_url:
        print("Error: MONGODB_URL not found in environment variables")
        print("Make sure you have a .env file with MONGODB_URL set")
        sys.exit(1)

    client = AsyncIOMotorClient(mongodb_url)
    db = client[database_name]

    # Get active season
    seasons_col = db["kvk_seasons"]
    active_season = await seasons_col.find_one({"status": "active"})

    if not active_season:
        print("Error: No active season found")
        sys.exit(1)

    season_id = active_season["season_id"]
    season_name = active_season.get("name", season_id)
    print(f"Found active season: {season_name} ({season_id})")

    # Get current data with players
    current_col = db["current_data"]
    current_data = await current_col.find_one({"kvk_season_id": season_id})

    if not current_data:
        print(f"Error: No player data found for season {season_id}")
        sys.exit(1)

    players = current_data.get("players", [])
    print(f"Found {len(players)} players")

    client.close()
    return players, season_name, season_id


def create_dkp_excel(players, season_name, season_id):
    """Create Excel file with DKP template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Final DKP"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for input
    formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green for calculated
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Rank",
        "Governor ID",
        "Governor Name",
        "Fight T4 Kills",
        "Fight T5 Kills",
        "T4 Deaths",
        "T5 Deaths",
        "Kill Score",
        "Death Score",
        "Total DKP"
    ]

    # Column widths
    col_widths = [8, 15, 25, 15, 15, 18, 18, 15, 15, 15]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sort players by fight KP for initial ordering
    sorted_players = sorted(
        players,
        key=lambda p: (p.get('fight_t4_kills', 0) * T4_KILL_WEIGHT +
                      p.get('fight_t5_kills', 0) * T5_KILL_WEIGHT),
        reverse=True
    )

    # Add player data
    for row, player in enumerate(sorted_players, 2):
        governor_id = player.get('governor_id', '')
        governor_name = player.get('governor_name', '')

        # Get fight period kills (or fallback to delta)
        fight_t4_kills = player.get('fight_t4_kills', 0)
        fight_t5_kills = player.get('fight_t5_kills', 0)

        # If no fight period kills, use delta
        if fight_t4_kills == 0 and fight_t5_kills == 0:
            delta = player.get('delta', {})
            fight_t4_kills = delta.get('t4_kills', 0)
            fight_t5_kills = delta.get('t5_kills', 0)

        # Get existing verified deaths if any
        verified = player.get('verified_deaths', {})
        t4_deaths = verified.get('t4_deaths', 0) if verified.get('verified') else 0
        t5_deaths = verified.get('t5_deaths', 0) if verified.get('verified') else 0

        # Row data
        ws.cell(row=row, column=1, value=row-1).alignment = center_align  # Rank (will update after sort)
        ws.cell(row=row, column=2, value=governor_id).alignment = center_align
        ws.cell(row=row, column=3, value=governor_name)
        ws.cell(row=row, column=4, value=fight_t4_kills).alignment = center_align
        ws.cell(row=row, column=5, value=fight_t5_kills).alignment = center_align

        # Death input cells (yellow background)
        t4_death_cell = ws.cell(row=row, column=6, value=t4_deaths)
        t4_death_cell.alignment = center_align
        t4_death_cell.fill = input_fill

        t5_death_cell = ws.cell(row=row, column=7, value=t5_deaths)
        t5_death_cell.alignment = center_align
        t5_death_cell.fill = input_fill

        # Formula cells (green background)
        # Kill Score = T4 Kills × 1 + T5 Kills × 2
        kill_score_cell = ws.cell(row=row, column=8)
        kill_score_cell.value = f"=D{row}*{T4_KILL_WEIGHT}+E{row}*{T5_KILL_WEIGHT}"
        kill_score_cell.alignment = center_align
        kill_score_cell.fill = formula_fill

        # Death Score = T4 Deaths × 4 + T5 Deaths × 8
        death_score_cell = ws.cell(row=row, column=9)
        death_score_cell.value = f"=F{row}*{T4_DEATH_WEIGHT}+G{row}*{T5_DEATH_WEIGHT}"
        death_score_cell.alignment = center_align
        death_score_cell.fill = formula_fill

        # Total DKP = Kill Score + Death Score
        total_cell = ws.cell(row=row, column=10)
        total_cell.value = f"=H{row}+I{row}"
        total_cell.alignment = center_align
        total_cell.fill = formula_fill
        total_cell.font = Font(bold=True)

        # Add borders to all cells
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    # Add info sheet
    info_ws = wb.create_sheet("Info")
    info_ws["A1"] = "Final DKP Calculator"
    info_ws["A1"].font = Font(bold=True, size=14)

    info_ws["A3"] = "Season:"
    info_ws["B3"] = season_name
    info_ws["A4"] = "Season ID:"
    info_ws["B4"] = season_id
    info_ws["A5"] = "Export Date:"
    info_ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    info_ws["A6"] = "Total Players:"
    info_ws["B6"] = len(players)

    info_ws["A8"] = "DKP Formula:"
    info_ws["A8"].font = Font(bold=True)
    info_ws["A9"] = "T4 Kills × 1"
    info_ws["A10"] = "T5 Kills × 2"
    info_ws["A11"] = "T4 Deaths × 4"
    info_ws["A12"] = "T5 Deaths × 8"

    info_ws["A14"] = "Instructions:"
    info_ws["A14"].font = Font(bold=True)
    info_ws["A15"] = "1. Yellow cells are for manual input (T4/T5 Deaths)"
    info_ws["A16"] = "2. Green cells are auto-calculated"
    info_ws["A17"] = "3. Enter deaths for each player, DKP will update automatically"
    info_ws["A18"] = "4. Sort by 'Total DKP' column (J) when done"

    info_ws.column_dimensions["A"].width = 20
    info_ws.column_dimensions["B"].width = 30

    # Save file
    filename = f"Final_DKP_{season_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"\n✅ Excel file created: {filename}")
    print(f"   - {len(players)} players exported")
    print(f"   - Yellow cells: Enter T4/T5 deaths")
    print(f"   - Green cells: Auto-calculated scores")
    print(f"\nOpen the file and enter deaths to calculate final DKP!")

    return filename


async def main():
    print("=" * 50)
    print("Final DKP Template Exporter")
    print("=" * 50)

    players, season_name, season_id = await get_player_data()
    create_dkp_excel(players, season_name, season_id)


if __name__ == "__main__":
    asyncio.run(main())
