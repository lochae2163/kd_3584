"""
Quick fix: Update baseline date for season_6 to correct date from Excel file.
"""
import asyncio
import os
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from datetime import datetime
import openpyxl

load_dotenv()

async def fix_baseline_date():
    # Extract correct date from baseline file
    excel_file = "../excel_files/kingdom_scan_baseline_data.xlsx"

    try:
        wb = openpyxl.load_workbook(excel_file)
        if 'Summary' not in wb.sheetnames:
            print("❌ No Summary sheet found")
            return

        ws = wb['Summary']
        date_cell = ws.cell(2, 6).value

        if not date_cell:
            print("❌ No date found in cell F2")
            return

        # Parse date
        if isinstance(date_cell, str) and 'UTC' in date_cell:
            date_str = date_cell.replace(' UTC', '').strip()
            correct_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
        elif isinstance(date_cell, datetime):
            correct_date = date_cell
        else:
            print(f"❌ Unexpected date format: {date_cell}")
            return

        print(f"✅ Extracted date from Excel: {correct_date}")

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return

    # Update database
    client = AsyncIOMotorClient(os.getenv("MONGODB_URL"))
    db = client["kvk_tracker"]
    baselines = db["baselines"]

    # Get current baseline
    baseline = await baselines.find_one({"kvk_season_id": "season_6"})
    if not baseline:
        print("❌ No baseline found for season_6")
        client.close()
        return

    print(f"\nCurrent baseline:")
    print(f"  File: {baseline.get('file_name')}")
    print(f"  Old timestamp: {baseline.get('timestamp')}")
    print(f"  New timestamp: {correct_date}")

    # Update
    result = await baselines.update_one(
        {"kvk_season_id": "season_6"},
        {"$set": {"timestamp": correct_date}}
    )

    if result.modified_count > 0:
        print(f"\n✅ Baseline date updated successfully!")
        print(f"   Changed from: {baseline.get('timestamp')}")
        print(f"   Changed to:   {correct_date}")
    else:
        print("\n⚠️  No changes made (date might already be correct)")

    client.close()

if __name__ == "__main__":
    print("=" * 70)
    print("FIX BASELINE DATE - Season 6")
    print("=" * 70)
    asyncio.run(fix_baseline_date())
    print("=" * 70)
