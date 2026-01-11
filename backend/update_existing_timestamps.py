"""
Update existing database records to use file dates from Excel files.

This script:
1. Finds all baseline, current_data, and upload_history records
2. For each record with a file_name ending in .xlsx/.xls
3. Looks for the Excel file in a specified directory
4. Extracts the file date from Summary sheet
5. Updates the timestamp in the database

Usage:
    python update_existing_timestamps.py --excel-dir /path/to/excel/files

    or

    python update_existing_timestamps.py --excel-dir /path/to/excel/files --dry-run
"""

import asyncio
import os
import sys
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from pathlib import Path
import argparse

# Load environment variables
load_dotenv()

MONGODB_URL = os.getenv("MONGODB_URL")

def extract_date_from_excel(excel_path: str):
    """Extract date from Excel Summary sheet (Row 2, Column F)."""
    try:
        wb = openpyxl.load_workbook(excel_path)
        if 'Summary' not in wb.sheetnames:
            print(f"  ⚠️  No Summary sheet found in {excel_path}")
            return None

        ws = wb['Summary']
        date_cell = ws.cell(2, 6).value  # Row 2, Column F

        if not date_cell:
            print(f"  ⚠️  No date in cell F2 in {excel_path}")
            return None

        # Parse date string like "2025-12-17 20:21 UTC"
        if isinstance(date_cell, str) and 'UTC' in date_cell:
            date_str = date_cell.replace(' UTC', '').strip()
            file_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
            print(f"  ✅ Extracted date: {file_date}")
            return file_date
        elif isinstance(date_cell, datetime):
            print(f"  ✅ Extracted date: {date_cell}")
            return date_cell
        else:
            print(f"  ⚠️  Unexpected date format: {date_cell}")
            return None

    except Exception as e:
        print(f"  ❌ Error reading {excel_path}: {e}")
        return None


async def update_timestamps(excel_dir: str, dry_run: bool = False):
    """Update timestamps in database for all Excel-based uploads."""

    # Connect to MongoDB
    client = AsyncIOMotorClient(MONGODB_URL)
    db = client["kvk_tracker"]

    print("=" * 70)
    print("DATABASE TIMESTAMP UPDATE SCRIPT")
    print("=" * 70)
    print(f"Excel directory: {excel_dir}")
    print(f"Mode: {'DRY RUN (no changes)' if dry_run else 'LIVE UPDATE'}")
    print("=" * 70)
    print()

    excel_dir_path = Path(excel_dir)
    if not excel_dir_path.exists():
        print(f"❌ Directory not found: {excel_dir}")
        return

    # Get all Excel files in directory
    excel_files = {}
    for ext in ['*.xlsx', '*.xls']:
        for filepath in excel_dir_path.glob(ext):
            excel_files[filepath.name] = filepath

    print(f"Found {len(excel_files)} Excel files in directory")
    print()

    # Track statistics
    stats = {
        'baselines_checked': 0,
        'baselines_updated': 0,
        'current_checked': 0,
        'current_updated': 0,
        'history_checked': 0,
        'history_updated': 0,
        'files_not_found': 0,
        'date_extraction_failed': 0
    }

    # Update baselines
    print("📊 Checking BASELINES collection...")
    print("-" * 70)
    baselines_col = db["baselines"]
    async for baseline in baselines_col.find({}):
        stats['baselines_checked'] += 1
        file_name = baseline.get('file_name', '')
        season_id = baseline.get('kvk_season_id', 'unknown')

        print(f"\n{stats['baselines_checked']}. Season: {season_id}, File: {file_name}")

        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            print("  ⏭️  Skipping (not an Excel file)")
            continue

        if file_name not in excel_files:
            print(f"  ⚠️  Excel file not found in directory")
            stats['files_not_found'] += 1
            continue

        excel_path = excel_files[file_name]
        file_date = extract_date_from_excel(str(excel_path))

        if not file_date:
            stats['date_extraction_failed'] += 1
            continue

        current_timestamp = baseline.get('timestamp')
        print(f"  Current timestamp: {current_timestamp}")
        print(f"  New timestamp: {file_date}")

        if not dry_run:
            await baselines_col.update_one(
                {"_id": baseline["_id"]},
                {"$set": {"timestamp": file_date}}
            )
            print("  ✅ Updated!")
        else:
            print("  🔍 Would update (dry run)")

        stats['baselines_updated'] += 1

    print()
    print("=" * 70)

    # Update current_data
    print("\n📊 Checking CURRENT_DATA collection...")
    print("-" * 70)
    current_col = db["current_data"]
    async for current in current_col.find({}):
        stats['current_checked'] += 1
        file_name = current.get('file_name', '')
        season_id = current.get('kvk_season_id', 'unknown')

        print(f"\n{stats['current_checked']}. Season: {season_id}, File: {file_name}")

        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            print("  ⏭️  Skipping (not an Excel file)")
            continue

        if file_name not in excel_files:
            print(f"  ⚠️  Excel file not found in directory")
            stats['files_not_found'] += 1
            continue

        excel_path = excel_files[file_name]
        file_date = extract_date_from_excel(str(excel_path))

        if not file_date:
            stats['date_extraction_failed'] += 1
            continue

        current_timestamp = current.get('timestamp')
        print(f"  Current timestamp: {current_timestamp}")
        print(f"  New timestamp: {file_date}")

        if not dry_run:
            await current_col.update_one(
                {"_id": current["_id"]},
                {"$set": {"timestamp": file_date}}
            )
            print("  ✅ Updated!")
        else:
            print("  🔍 Would update (dry run)")

        stats['current_updated'] += 1

    print()
    print("=" * 70)

    # Update upload_history
    print("\n📊 Checking UPLOAD_HISTORY collection...")
    print("-" * 70)
    history_col = db["upload_history"]
    async for history in history_col.find({}):
        stats['history_checked'] += 1
        file_name = history.get('file_name', '')
        season_id = history.get('kvk_season_id', 'unknown')

        print(f"\n{stats['history_checked']}. Season: {season_id}, File: {file_name}")

        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            print("  ⏭️  Skipping (not an Excel file)")
            continue

        if file_name not in excel_files:
            print(f"  ⚠️  Excel file not found in directory")
            stats['files_not_found'] += 1
            continue

        excel_path = excel_files[file_name]
        file_date = extract_date_from_excel(str(excel_path))

        if not file_date:
            stats['date_extraction_failed'] += 1
            continue

        current_timestamp = history.get('timestamp')
        print(f"  Current timestamp: {current_timestamp}")
        print(f"  New timestamp: {file_date}")

        if not dry_run:
            await history_col.update_one(
                {"_id": history["_id"]},
                {"$set": {"timestamp": file_date}}
            )
            print("  ✅ Updated!")
        else:
            print("  🔍 Would update (dry run)")

        stats['history_updated'] += 1

    print()
    print("=" * 70)
    print("\n📈 SUMMARY")
    print("=" * 70)
    print(f"Baselines checked: {stats['baselines_checked']}")
    print(f"Baselines updated: {stats['baselines_updated']}")
    print()
    print(f"Current data checked: {stats['current_checked']}")
    print(f"Current data updated: {stats['current_updated']}")
    print()
    print(f"History records checked: {stats['history_checked']}")
    print(f"History records updated: {stats['history_updated']}")
    print()
    print(f"⚠️  Files not found: {stats['files_not_found']}")
    print(f"❌ Date extraction failed: {stats['date_extraction_failed']}")
    print("=" * 70)

    if dry_run:
        print("\n🔍 This was a DRY RUN - no changes were made")
        print("Run without --dry-run to actually update the database")
    else:
        print("\n✅ Database update complete!")

    client.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Update database timestamps from Excel files')
    parser.add_argument('--excel-dir', required=True, help='Directory containing Excel files')
    parser.add_argument('--dry-run', action='store_true', help='Preview changes without updating database')

    args = parser.parse_args()

    asyncio.run(update_timestamps(args.excel_dir, args.dry_run))
