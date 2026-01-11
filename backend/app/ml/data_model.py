"""
ML-powered Data Processing Model for KvK Tracker
================================================
This module handles:
1. CSV data cleaning and validation
2. Delta calculation between baseline and current data
3. Data transformation for frontend display
4. Anomaly detection (optional future feature)

Author: Kingdom 3584
Project: KvK Stats Tracker (Data Science Portfolio)
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from io import StringIO, BytesIO
from dataclasses import dataclass
import logging
import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class PlayerStats:
    """Data class for player statistics."""
    governor_id: str
    governor_name: str
    power: int
    kill_points: int
    deads: int
    t4_kills: int
    t5_kills: int


@dataclass
class PlayerDelta:
    """Data class for player stats with delta from baseline."""
    governor_id: str
    governor_name: str
    power: int
    kill_points: int
    deads: int
    t4_kills: int
    t5_kills: int
    delta_power: int
    delta_kill_points: int
    delta_deads: int
    delta_t4_kills: int
    delta_t5_kills: int


class KvKDataModel:
    """
    Machine Learning Model for KvK Data Processing
    
    This class handles all data processing tasks:
    - CSV parsing and cleaning
    - Data validation
    - Delta calculation
    - Statistical analysis
    
    Usage:
        model = KvKDataModel()
        clean_data = model.process_csv(csv_content)
        deltas = model.calculate_deltas(baseline_data, current_data)
    """
    
    # Expected columns in CSV
    REQUIRED_COLUMNS = [
        'governor_id',
        'governor_name', 
        'power',
        'deads',
        'kill_points',
        't4_kills',
        't5_kills'
    ]
    
    # Numeric columns that need cleaning
    NUMERIC_COLUMNS = ['power', 'deads', 'kill_points', 't4_kills', 't5_kills']
    
    def __init__(self):
        """Initialize the data model."""
        self.baseline_data: Optional[pd.DataFrame] = None
        self.current_data: Optional[pd.DataFrame] = None
        self.processing_stats = {}
    
    # ==========================================
    # Data Cleaning Methods
    # ==========================================
    
    def clean_numeric_value(self, value) -> int:
        """
        Clean a numeric value by removing commas and converting to int.
        
        Handles:
        - "230,639,240" -> 230639240
        - "230639240" -> 230639240
        - 230639240 -> 230639240
        - NaN/None -> 0
        
        Args:
            value: The value to clean
            
        Returns:
            int: Cleaned integer value
        """
        if pd.isna(value) or value is None:
            return 0
        
        if isinstance(value, (int, float)):
            return int(value)
        
        # Remove commas, quotes, and whitespace
        cleaned = str(value).replace(',', '').replace('"', '').replace("'", "").strip()
        
        if not cleaned:
            return 0
        
        try:
            return int(float(cleaned))
        except (ValueError, TypeError):
            logger.warning(f"Could not convert value: {value}")
            return 0
    
    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean entire dataframe.
        
        Operations:
        1. Strip whitespace from column names
        2. Convert numeric columns
        3. Convert governor_id to string
        4. Remove duplicate entries
        5. Handle missing values
        
        Args:
            df: Raw dataframe
            
        Returns:
            pd.DataFrame: Cleaned dataframe
        """
        # Make a copy to avoid modifying original
        df = df.copy()
        
        # Clean column names (strip whitespace, lowercase)
        df.columns = df.columns.str.strip().str.lower()
        
        # Clean numeric columns
        for col in self.NUMERIC_COLUMNS:
            if col in df.columns:
                df[col] = df[col].apply(self.clean_numeric_value)
        
        # Ensure governor_id is string
        if 'governor_id' in df.columns:
            df['governor_id'] = df['governor_id'].astype(str).str.strip()
        
        # Ensure governor_name is string
        if 'governor_name' in df.columns:
            df['governor_name'] = df['governor_name'].astype(str).str.strip()
        
        # Remove duplicates (keep last occurrence)
        df = df.drop_duplicates(subset=['governor_id'], keep='last')
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    # ==========================================
    # CSV Processing Methods
    # ==========================================
    
    def validate_csv(self, csv_content: str) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        """
        Validate CSV content and structure.
        
        Args:
            csv_content: Raw CSV string
            
        Returns:
            Tuple of (is_valid, message, dataframe or None)
        """
        try:
            # Try to read CSV
            df = pd.read_csv(StringIO(csv_content))
            
            # Check if empty
            if df.empty:
                return False, "CSV file is empty", None
            
            # Clean column names for comparison
            df.columns = df.columns.str.strip().str.lower()
            
            # Check required columns
            missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
            if missing_cols:
                return False, f"Missing required columns: {missing_cols}", None
            
            return True, "CSV is valid", df
            
        except Exception as e:
            return False, f"Failed to parse CSV: {str(e)}", None
    
    def process_csv(self, csv_content: str) -> Dict:
        """
        Process CSV content through the ML pipeline.

        Pipeline:
        1. Validate CSV structure
        2. Clean data
        3. Calculate statistics
        4. Return processed data

        Args:
            csv_content: Raw CSV string

        Returns:
            Dict with processed data and metadata
        """
        # Validate
        is_valid, message, df = self.validate_csv(csv_content)

        if not is_valid:
            return {
                "success": False,
                "error": message,
                "data": None
            }

        # Clean data
        df_cleaned = self.clean_dataframe(df)

        # Calculate processing stats
        self.processing_stats = {
            "rows_processed": len(df_cleaned),
            "columns": list(df_cleaned.columns),
            "timestamp": datetime.utcnow().isoformat()
        }

        # Convert to list of dicts for API response
        players = []
        for _, row in df_cleaned.iterrows():
            players.append({
                "governor_id": str(row['governor_id']),
                "governor_name": str(row['governor_name']),
                "stats": {
                    "power": int(row['power']),
                    "kill_points": int(row['kill_points']),
                    "deads": int(row['deads']),
                    "t4_kills": int(row['t4_kills']),
                    "t5_kills": int(row['t5_kills'])
                }
            })

        return {
            "success": True,
            "message": f"Successfully processed {len(players)} players",
            "player_count": len(players),
            "players": players,
            "processing_stats": self.processing_stats
        }

    def process_excel(self, excel_bytes: bytes, kingdom_id: str = "3584") -> Dict:
        """
        Process Excel file from Hero Scrolls Kingdom Scanner.

        This method:
        1. Auto-detects the correct sheet (kingdom_id or "Rolled Up {kingdom_id}")
        2. Extracts file date from Summary sheet
        3. Maps Excel columns to required CSV format
        4. Cleans and validates data
        5. Returns processed player data with file date

        Args:
            excel_bytes: Raw Excel file bytes
            kingdom_id: Kingdom number to look for (default "3584")

        Returns:
            Dict with processed data, metadata, and file_date
        """
        try:
            # Read Excel file
            excel_file = BytesIO(excel_bytes)

            # Extract file date from Summary sheet (Row 2, Column F)
            file_date = None
            try:
                wb = openpyxl.load_workbook(BytesIO(excel_bytes))
                if 'Summary' in wb.sheetnames:
                    ws = wb['Summary']
                    date_cell = ws.cell(2, 6).value  # Row 2, Column F (6)
                    if date_cell:
                        # Parse date string like "2025-12-17 20:21 UTC"
                        if isinstance(date_cell, str) and 'UTC' in date_cell:
                            # Remove ' UTC' and parse
                            date_str = date_cell.replace(' UTC', '').strip()
                            file_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M').isoformat()
                            logger.info(f"Extracted file date from Summary sheet: {file_date}")
                        elif isinstance(date_cell, datetime):
                            file_date = date_cell.isoformat()
                            logger.info(f"Extracted file date from Summary sheet: {file_date}")
            except Exception as e:
                logger.warning(f"Could not extract file date from Summary sheet: {e}")
                # If we can't extract date, use upload time as fallback
                file_date = None

            # Get all sheet names
            xls = pd.ExcelFile(excel_file)
            sheet_names = xls.sheet_names

            logger.info(f"Found sheets: {sheet_names}")

            # Auto-detect correct sheet
            # Priority: exact match > "Rolled Up" > first sheet with kingdom_id in name
            target_sheet = None

            # 1. Try exact match
            if kingdom_id in sheet_names:
                target_sheet = kingdom_id
            # 2. Try "Rolled Up {kingdom_id}"
            elif f"Rolled Up {kingdom_id}" in sheet_names:
                target_sheet = f"Rolled Up {kingdom_id}"
            # 3. Try any sheet containing kingdom_id
            else:
                for sheet in sheet_names:
                    if kingdom_id in sheet:
                        target_sheet = sheet
                        break

            # 4. Default to first non-summary sheet
            if not target_sheet:
                for sheet in sheet_names:
                    if sheet.lower() not in ['summary', 'top 10s', 'top10s']:
                        target_sheet = sheet
                        break

            if not target_sheet:
                return {
                    "success": False,
                    "error": f"Could not find data sheet for kingdom {kingdom_id}. Available sheets: {sheet_names}",
                    "data": None
                }

            logger.info(f"Using sheet: {target_sheet}")

            # Read the target sheet
            df = pd.read_excel(excel_file, sheet_name=target_sheet)

            # Check if empty
            if df.empty:
                return {
                    "success": False,
                    "error": f"Sheet '{target_sheet}' is empty",
                    "data": None
                }

            # Map Excel columns to required format
            # Hero Scrolls Excel columns (based on analysis):
            column_mapping = {
                'Governor ID': 'governor_id',
                'Governor Name': 'governor_name',
                'Power': 'power',
                'Deads': 'deads',
                'Kill Points': 'kill_points',
                'T4 Kills': 't4_kills',
                'T5 Kills': 't5_kills'
            }

            # Rename columns
            df = df.rename(columns=column_mapping)

            # Clean column names (strip whitespace, lowercase)
            df.columns = df.columns.str.strip().str.lower()

            # Check required columns exist
            missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
            if missing_cols:
                return {
                    "success": False,
                    "error": f"Missing required columns after mapping: {missing_cols}. Available columns: {list(df.columns)}",
                    "data": None
                }

            # Keep only required columns
            df = df[self.REQUIRED_COLUMNS]

            # Clean data
            df_cleaned = self.clean_dataframe(df)

            # Calculate processing stats
            self.processing_stats = {
                "rows_processed": len(df_cleaned),
                "columns": list(df_cleaned.columns),
                "sheet_used": target_sheet,
                "timestamp": datetime.utcnow().isoformat()
            }

            # Convert to list of dicts for API response
            players = []
            for _, row in df_cleaned.iterrows():
                players.append({
                    "governor_id": str(row['governor_id']),
                    "governor_name": str(row['governor_name']),
                    "stats": {
                        "power": int(row['power']),
                        "kill_points": int(row['kill_points']),
                        "deads": int(row['deads']),
                        "t4_kills": int(row['t4_kills']),
                        "t5_kills": int(row['t5_kills'])
                    }
                })

            return {
                "success": True,
                "message": f"Successfully processed {len(players)} players from sheet '{target_sheet}'",
                "player_count": len(players),
                "players": players,
                "processing_stats": self.processing_stats,
                "file_date": file_date  # Date from Excel file, not upload time
            }

        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}")
            return {
                "success": False,
                "error": f"Failed to process Excel file: {str(e)}",
                "data": None
            }
    
    # ==========================================
    # Delta Calculation Methods
    # ==========================================
    
    def calculate_player_delta(
        self, 
        baseline_stats: Dict, 
        current_stats: Dict
    ) -> Dict:
        """
        Calculate delta between baseline and current stats for one player.
        
        Args:
            baseline_stats: Player's baseline stats dict
            current_stats: Player's current stats dict
            
        Returns:
            Dict with delta values
        """
        delta = {}
        
        for field in self.NUMERIC_COLUMNS:
            baseline_val = baseline_stats.get(field, 0)
            current_val = current_stats.get(field, 0)
            delta[field] = current_val - baseline_val
        
        return delta
    
    def calculate_all_deltas(
        self,
        baseline_players: List[Dict],
        current_players: List[Dict]
    ) -> List[Dict]:
        """
        Calculate deltas for all players.

        New players (not in baseline) will have their current stats
        set as their new baseline with zero deltas.

        Args:
            baseline_players: List of player dicts from baseline
            current_players: List of player dicts from current snapshot

        Returns:
            List of player dicts with delta information
        """
        # Create baseline lookup by governor_id
        baseline_lookup = {}
        for player in baseline_players:
            gov_id = player.get('governor_id')
            baseline_lookup[gov_id] = player.get('stats', {})

        # Calculate deltas for each current player
        results = []
        new_players_added = []

        for player in current_players:
            gov_id = player.get('governor_id')
            current_stats = player.get('stats', {})

            # Check if player has any KvK-relevant stats (kill_points, deads, t4_kills, t5_kills)
            # Power alone doesn't count as they might have migrated out
            kvk_stats_fields = ['kill_points', 'deads', 't4_kills', 't5_kills']
            has_any_stats = any(current_stats.get(field, 0) > 0 for field in kvk_stats_fields)

            # Check if player exists in baseline
            if gov_id in baseline_lookup:
                baseline_stats = baseline_lookup[gov_id]

                if has_any_stats:
                    # Existing player with stats - calculate delta normally
                    delta = self.calculate_player_delta(baseline_stats, current_stats)
                    stats_to_show = current_stats
                    in_baseline = True
                    newly_added = False
                else:
                    # Player was in baseline but now has 0 KvK stats (migrated out)
                    # Keep showing their baseline stats, but with 0 delta
                    delta = {field: 0 for field in self.NUMERIC_COLUMNS}
                    stats_to_show = baseline_stats  # Show baseline stats, not current zeros
                    in_baseline = True
                    newly_added = False
                    logger.info(f"Player migrated out: {player.get('governor_name')} (ID: {gov_id}) - Showing baseline stats with 0 delta")
            else:
                # New player - use current stats as new baseline
                baseline_stats = current_stats
                delta = {field: 0 for field in self.NUMERIC_COLUMNS}
                stats_to_show = current_stats
                in_baseline = False
                newly_added = True

                # Track new player to add to baseline
                new_players_added.append({
                    "governor_id": gov_id,
                    "governor_name": player.get('governor_name'),
                    "stats": current_stats
                })
                logger.info(f"New player detected: {player.get('governor_name')} (ID: {gov_id}) - Setting as new baseline")

            # Add to results
            results.append({
                "governor_id": gov_id,
                "governor_name": player.get('governor_name'),
                "stats": stats_to_show,
                "delta": delta,
                "in_baseline": in_baseline,
                "newly_added_to_baseline": newly_added
            })

        # Log summary if new players were added
        if new_players_added:
            logger.info(f"Added {len(new_players_added)} new players to baseline automatically")

        return results
    
    # ==========================================
    # Statistical Analysis Methods
    # ==========================================
    
    def calculate_summary_stats(self, players: List[Dict]) -> Dict:
        """
        Calculate summary statistics for a list of players.

        For top_players:
        - kill_points: Uses delta (gained) - who gained most KP
        - t5_kills: Uses delta (gained) - who gained most T5 kills
        - t4_kills: Uses delta (gained) - who gained most T4 kills
        - deads: Uses delta (gained) - who gained most deaths
        - power: Uses total stats (overall power) - who has highest power

        Args:
            players: List of player dicts (with stats and delta)

        Returns:
            Dict with summary statistics
        """
        if not players:
            return {}

        # Convert to dataframe for easy stats calculation
        stats_list = [p.get('stats', {}) for p in players]
        df_stats = pd.DataFrame(stats_list)

        # Also get delta dataframe for specific top players
        delta_list = [p.get('delta', {}) for p in players]
        df_delta = pd.DataFrame(delta_list)

        summary = {
            "player_count": len(players),
            "totals": {},
            "averages": {},
            "top_players": {}
        }

        # Calculate totals and averages (always use stats)
        for col in self.NUMERIC_COLUMNS:
            if col in df_stats.columns:
                summary["totals"][col] = int(df_stats[col].sum())
                summary["averages"][col] = int(df_stats[col].mean())

        # Find top players - use delta for kill_points, t5_kills, t4_kills, deads
        # Use stats for power only
        for col in self.NUMERIC_COLUMNS:
            # Use delta for kill_points, t5_kills, t4_kills, deads
            if col in ['kill_points', 't5_kills', 't4_kills', 'deads']:
                if col in df_delta.columns:
                    idx = df_delta[col].idxmax()
                    summary["top_players"][col] = {
                        "name": players[idx].get('governor_name'),
                        "governor_id": players[idx].get('governor_id'),
                        "value": int(df_delta.loc[idx, col])  # Delta value (gained)
                    }
            # Use stats for power only
            else:
                if col in df_stats.columns:
                    idx = df_stats[col].idxmax()
                    summary["top_players"][col] = {
                        "name": players[idx].get('governor_name'),
                        "governor_id": players[idx].get('governor_id'),
                        "value": int(df_stats.loc[idx, col])  # Stats value (total)
                    }

        return summary
    
    def rank_players(
        self,
        players: List[Dict],
        sort_by: str = "kill_points",
        ascending: bool = False
    ) -> List[Dict]:
        """
        Rank players by a specific stat.

        Args:
            players: List of player dicts
            sort_by: Field to sort by (supports both stats and delta fields)
            ascending: Sort order

        Returns:
            List of player dicts with rank added
        """
        if not players:
            return []

        # Determine if sorting by delta (gained) fields
        # Fields ending in _gained are from delta object
        if sort_by.endswith('_gained'):
            # Extract the base field name (e.g., kill_points_gained -> kill_points)
            base_field = sort_by.replace('_gained', '')
            sorted_players = sorted(
                players,
                key=lambda p: p.get('delta', {}).get(base_field, 0),
                reverse=not ascending
            )
        else:
            # Sort by stats field
            sorted_players = sorted(
                players,
                key=lambda p: p.get('stats', {}).get(sort_by, 0),
                reverse=not ascending
            )

        # Add rank
        for i, player in enumerate(sorted_players, 1):
            player['rank'] = i

        return sorted_players


# Create singleton instance
kvk_model = KvKDataModel()